###Searchable student analytics db built by Steve Reitz
##Python 3.6 using openpyxl module
##
##
##
##
## TODO:
# 4. Simple searches = saved?
## 1. Add searching by state
# 2. Daily Progress Report
# 5. Make it searchable by name too
# 6. Add multiple searching capabilities
# 3.. Gradebook functionality
# 7. CSV reading functionality

import os, openpyxl, pprint
from openpyxl import Workbook
##Master sheet
#wbTest= openpyxl.load_workbook('test.xlsx')
#sheet= wbTest.get_sheet_by_name('Complete')
##Evaluation sheet
wbCompletion= openpyxl.load_workbook('test2.xlsx')
sheet2= wbCompletion.get_sheet_by_name('Complete')
##NLI Sheet.
##NOTE: You will always have to save it as a xlsv as when you download it from BQ, it's a csv.
wbNLI=openpyxl.load_workbook('test3.xlsx')
sheet3= wbNLI.get_sheet_by_name('Students-Never-Logged-into-Blue')
##Master Sheet with all of the states
wbStates=openpyxl.load_workbook('test4.xlsx')
sheet= wbStates.get_sheet_by_name('Master')
##Dictionaries that store everything
student={}
userNLI={}
emailCheck={}
completeCount=0
loggedInCount=0

#test
#test2

##This next loop fills the dictionaries
for row in range(1, 427):
        user= sheet3['D'+str(row)].value
        userNLI.setdefault(user,{})
        email= sheet2['C'+str(row)].value
        evalEmail=sheet2['F'+str(row)].value
        emailCheck.update({email: evalEmail})
        
##This next loop fills out the sheet ala Google sheets but without any weird sorting, b/c yay Python
sheet=0
for sheet in wbStates:
        n=0
        wbStates.sheetnames[n]
        #print(wbStates.active)
        loggedInCount=0
        completeCount=0
        for rowNum in range(1,427):
            allUser= sheet['E'+str(rowNum)].value
            allEmails= sheet['D'+str(rowNum)].value
            allFirstNames= sheet['A'+str(rowNum)].value
            allLastNames= sheet['B'+str(rowNum)].value
            allAgencies= sheet.title
            #Sets up the searchable categories
            student.setdefault(allUser,{"First Name": allFirstNames,"Last Name": allLastNames,"Agency": allAgencies, "User": allUser, "Email": allEmails, "¿Ha accedido al curso?": '',"¿Completó el curso?":''})
            yesOrNo= sheet.cell(row=rowNum, column=7)
            if allUser in userNLI:
                yesOrNo.value= 'No'
                student[allUser].update({"¿Ha accedido al curso?":'No'})
            else:
                yesOrNo.value= 'Sí'
                student[allUser].update({"¿Ha accedido al curso?":'Sí'})
                loggedInCount=loggedInCount+1
            if allEmails in emailCheck:
                sheet.cell(row=rowNum, column=8).value= 'Completó'
                student[allUser].update({"¿Completó el curso?":'Completó'})
                completeCount=completeCount+1
                #print(completeCount)
            elif( yesOrNo.value == 'Sí'):
                sheet.cell(row=rowNum, column=8).value= 'En progreso'
                student[allUser].update({"¿Completó el curso?":'En progreso'})                      
            else:
                sheet.cell(row=rowNum, column=8).value= 'No ha iniciado sesión'
                student[allUser].update({"¿Completó el curso?":'No ha iniciado sesión'})

##         IGNORE THIS CODE. WIP THAT IS LOW PRIORITY SILLINESS
##                
##        print(str(sheet.max_row)+sheet.title)
##        sheet['G'+str(sheet.max_row)].value= loggedInCount
##        loggedInPercent= (loggedInCount/427)*100
##        #sheet[''+sheet.max_row].value=str(loggedInPercent) + "% have logged into the course"
##        #print(str(loggedInPercent) +"% have logged into the course")
##        print(loggedInCount)
##        sheet['H'+str(sheet.max_row)].value= completeCount
##        completePercent= (completeCount/427)*100
        #sheet['H429'].value= str(completePercent) +"% have completed the course"
##        print(str(completePercent)+"% have completed the course")
##        sheets= wbStates.sheetnames
##        for s in sheets:
##                if s!=sheet.title:
##                        sheet_name= wbStates.get_sheet_by_name(s)
##                        wbStates.remove_sheet(sheet_name)
##        wbStates.save(sheet.title+'.xlsx')
##        n=n+1
##        wbStates= openpyxl.load_workbook('test4.xlsx')

#Searches for 
def saveUserSearch(saveName, dictTerm, dictTerm2):
        print("Would you like to save this list?")
        yesOrNo=input()
        if yesOrNo=="Yes":
                wbNew= Workbook()
                ws= wbNew.active
                ws.title= str(saveName)
                for k,v in student[dictTerm2].items():
                        for row in range(1,(len(student[dictTerm2])+1)):
                                row2=row-1
                                ws['A'+str(row)].value= list(student[dictTerm2].keys())[row2]
                                ws['B'+str(row)].value= list(student[dictTerm2].values())[row2]
        wbNew.save(str(saveName)+'.xlsx')
        print("Saving file "+str(saveName)+'.xlsx')

 #Searches for unsent evaluations               
def saveUnsentEvals(saveName, dictTerm):
        print("Would you like to save this list?")
        yesOrNo=input()
        wbNew= Workbook()
        ws= wbNew.active
        ws.title= str(saveName)
        if yesOrNo=="Yes":
                for k,v in emailCheck.items():
                        for row in range(1,(len(emailCheck)+1)):
                                if values==None:
                                        row2=row-1
                                        ws['A'+str(row)].value= list(emailCheck)[row2]
                                        ws['B'+str(row)].value= list(emailCheck)[row2]
                wbNew.save(str(saveName)+'.xlsx')
                print("Saving file "+str(saveName)+'.xlsx')
        else:
                print("")
                
##These next three func
def agencyCompletionResults(agencyName):
        ##DESC: 
        ##TODO: Completion,
        ##TODO:login, eval filters
        print("Got here")
##        if agencyName in list(student.:
##                        print(k)
        
def agencyLoginResults(agencyName):
        print(a)
def agencyEvalResults(agencyName):
        print(a)

def generalSearch():
##DESC: This next part is the bit that allows you to search for users. Aka the functionality that any LMS should have        
        print("What would you like to know? Would you like to search a user? Would you like to see who hasn't received an evaluation?")
        searchSpec=input()
        if 'user' in searchSpec:
                print('What username would you like to search for?')
                userSearch=input()
                if userSearch in student.keys():
                    print("What would you like to know? \n You can answer with \n Everything, Usename, Eval sent?, Email, Logged In, Complete, or Agency")
                    specify=input()
                    if(specify=='Everything'):
                        pprint.pprint(student[userSearch])
                        saveUserSearch(str(userSearch)+ " User Info", userSearch, str(userSearch))
                    elif(specify=='Username'):
                        print(student[userSearch]['User'])
                    elif(specify=='Email'):
                        print(student[userSearch]['Email'])
                        #saveSpecificUserSearch(str(userSearch)+ " Username", userSearch, str(userSearch),'Email')
                    elif(specify=='Logged In'):
                        print(student[userSearch]['¿Ha accedido al curso?'])
                        #saveSpecificUserSearch(str(userSearch)+ " Username", userSearch, str(userSearch),'¿Ha accedido al curso?')
                    elif(specify=='Complete'):
                        print(student[userSearch]['¿Completó el curso?'])
                    elif(specify=='Agency'):
                        print(student[userSearch]['Agency'])
                    elif(specify=='Eval sent?'):
                        if(emailCheck[(student[userSearch]['Email'])] is None):
                           print('No')
                        else:
                           print(emailCheck[(student[userSearch]['Email'])])
                else:
                    print("Sorry, user not found.")
        elif 'eval' in searchSpec:
                wbNew= Workbook()
                ws= wbNew.active
                for k,v in emailCheck.items():
                        if v==None:
                            print(k)
                numRow=1
                for row in range(1,(len(emailCheck)+1)):        
                        row2=(row-1)
                        if list(emailCheck.values())[row2]==None:
                                ws['A'+str(numRow)].value= list(emailCheck.keys())[row2]
                                ws['B'+str(numRow)].value= list(emailCheck.values())[row2]
                                numRow=numRow+1
                        else:
                                 row=row+1
                wbNew.save(' No Evals sent.xlsx')
                print("This list can also be found as  "+'No Evals sent.xlsx in ' +os.getcwd())
                                    #saveUnsentEvals(" Evals Not Sent Yet", None)
        elif 'agency' in searchSpec:
                print('What agency would you like to search for?')
                agencySearch=input()
                print('Would you like to filter search results for this agency?\n By completion? By logged in? By evaluation sent?')
                filterSearch=input()
                if(filterSearch=="Completion"):
                        agencyCompletionResults(agencySearch)

generalSearch()
##Saves it as a Excel sheet
print('Saved updated.xlsx')
wbStates.save('updated.xlsx')
