import openpyxl

wb1 = openpyxl.load_workbook('F:/Steve/Final Program Enrollments.xlsx')
wb2= openpyxl.load_workbook('F:/Steve/Final Data Dump Users.xlsx')
wb3= openpyxl.load_workbook('F:/Steve/Roster for Salesforce.xlsx')


ws1 = wb1.active
dataDump = wb2.active
salesforceWS= wb3.active
from openpyxl.utils import range_boundaries
min_col, min_row, max_col, max_row = range_boundaries('A1:H')

#This looks at colunn B
column= 1
counter=0

for programEnrollments in ws1.iter_rows():
        for rowNum in range(len(dataDump['A'])):
            if programEnrollments[column].value== dataDump['A'][rowNum].value:
                programEnrollments[column+12].value=dataDump['S'][rowNum].value
                programEnrollments[column+14].value=dataDump['C'][rowNum].value
        for sfID in range(len(salesforceWS['G'])):
            if programEnrollments[column+14].value.lower() in salesforceWS['E'][sfID].value.lower():
                programEnrollments[column+13].value=salesforceWS['G'][sfID].value                        
                counter=counter+1
        ##                if y==295:
        ##                    wb1.save("file3.xlsx")
        ##                    break
        ##                    exit 
                print("appending..."+str(counter))

wb1.save("F:/Steve/file3.xlsx")
