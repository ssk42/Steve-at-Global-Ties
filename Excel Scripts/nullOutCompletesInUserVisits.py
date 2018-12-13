import openpyxl

wb1 = openpyxl.load_workbook('F:/Steve/user_visits.xlsx')
wb2= openpyxl.load_workbook('F:/Steve/wbusers.xlsx')
wb3= openpyxl.load_workbook('F:/Steve/Finished participants Fall 2018.xlsx')


ws1 = wb1.active
bridgeUsersWS = wb2.active
finishersWS= wb3.active
from openpyxl.utils import range_boundaries
min_col, min_row, max_col, max_row = range_boundaries('A1:H')

#This looks at colunn B
column= 2
counter=0
#test

##TODO:
#Convert to Email
#Check Email is Complete
#If not, Make Email=0



for userVisits in ws1.iter_rows():
        if userVisits[column].value!=2:
                for rowNum in range(len(bridgeUsersWS['A'])):
                    if userVisits[column].value== bridgeUsersWS['A'][rowNum].value:
                        userVisits[column].value=bridgeUsersWS['B'][rowNum].value
##                for finisher in range(len(finishersWS['E'])):
##                    if userVisits[column].value== finishersWS['E'][finisher].value.lower():
##                        userVisits[column].value=finishersWS['C'][finisher].value.lower()                        
                       counter=counter+1
##                        #test
                        
                        print("appending..."+str(counter))

wb1.save("F:/Steve/testingFile.xlsx")
