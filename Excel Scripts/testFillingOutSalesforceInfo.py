import openpyxl

wb1 = openpyxl.load_workbook('C:/Users/sreitz/Downloads/Master Roster For Bridge_TEMP.xlsx')
wb3= openpyxl.load_workbook('F:/Steve/Roster for Salesforce.xlsx')

ws1 = wb1['Unenrolled on 11.19.18']
ws2 = wb3.active
from openpyxl.utils import range_boundaries
min_col, min_row, max_col, max_row = range_boundaries('A1:H')

check= 4
y=0

for row in ws1.iter_rows():
        for x in range(len(ws2['E'])):
            #print(row[check].value)
            #print(ws2['E'][x].value)
            if row[check].value.lower() in ws2['E'][x].value.lower():
                row[check+3].value=ws2['G'][x].value
                y=y+1
                if y==295:
                    wb1.save("file3.xlsx")
                    break
                    exit 
                print("appending..."+str(y))

wb1.save("F:/Steve/Steve-at-Global-Ties/forGithub.xlsx")
