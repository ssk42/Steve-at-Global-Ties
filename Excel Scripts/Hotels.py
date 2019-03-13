import pandas as pd
from openpyxl.reader.excel import load_workbook
import openpyxl, numpy
from datetime import timedelta

counter=1

##This is the SurveyGizmo export. This is the master data.
hotelDF= pd.read_csv('F:/Steve/hotels.csv', encoding='latin1')

## These identify the proper columns from the original hotel spreadsheet.
blackout= hotelDF.columns[hotelDF.columns.str.startswith('Blackout')]
oneIVLP= hotelDF.columns[hotelDF.columns.str.startswith('Hotel can accommodate ONE (1)')]
twoIVLP= hotelDF.columns[hotelDF.columns.str.startswith('Hotel CAN accommodate two (2)')]
hotelNamesDF= pd.DataFrame(hotelDF['Hotel Name'])
wb= load_workbook('F:/Steve/Hotel Date Ranges.xlsx')
ws=wb.active
wbProjects= load_workbook('F:/Steve/Programs.xlsx')
wsProjects= wbProjects.active

## The following section makes the newly found columns into dataFrames. It starts with Blackout dates, then One IVLP Group, then Two IVLP Groups
hotelBlackoutDF= hotelDF[blackout]
hotelBlackoutDF= pd.merge(hotelNamesDF,hotelBlackoutDF, on=hotelBlackoutDF.index)
# hotelBlackoutDF.to_csv('F:/Steve/testBlackoutHotels.csv')
hotelOneDF= hotelDF[oneIVLP]
hotelOneDF= pd.merge(hotelNamesDF,hotelOneDF, on=hotelOneDF.index)
# hotelOneDF.to_csv('F:/Steve/testOneHotels.csv')
hotelTwoDF= hotelDF[twoIVLP]
hotelTwoDF= pd.merge(hotelNamesDF,hotelTwoDF, on=hotelTwoDF.index)
# hotelTwoDF.to_csv('F:/Steve/testTwoHotels.csv')

def hotelName(num):
	if num==0:
		return 'Embassy Row Hotel'
	if num==1:
		return 'Hampton Inn DC Convention Center'
	if num==2:
		return 'The Churchill Hotel'
	if num==3:
		return 'The Darcy'
	if num==4:
		return 'Residence Inn Washington DC Downtown'
	if num==5:
		return 'Homewood Suites by Hilton, Downtown'
	if num==6:
		return 'Washington Marriott Georgetown'
	if num==7:
		return 'ONE WASHINGTON CIRCLE HOTEL'
	if num==8:
		return 'The Fairfax At Embassy Row'
	if num==9:
		return 'Washington Plaza Hotel'
	if num==10:
		return 'Kimpton Hotel Palomar'
	if num==11:
		return 'Kimpton Carlyle Hotel'
	if num==12:
		return 'The Donovan'
	if num==13:
		return 'Washington Hilton'
	if num==14:
		return 'Club Quarters Washington DC'
	if num==15:
		return 'Melrose'

## The following section starts the main gist of the program. The first part sets up the dictionary entries per unique IVLP date range.
dateRangesDF= pd.read_csv('F:/Steve/Hotel Date Ranges.csv')
dateRangesDF= pd.DataFrame(dateRangesDF['Check-In'])

projectCounter=2
def projectAndHotel(hotelDict,counter,projectCounter,groupType):
	for row_cells in wsProjects.iter_rows():
			for cell in row_cells:
				if pd.notnull(ws[counter][0].value):
					if ws[counter][0].value+timedelta(days=2)==cell.value or ws[counter][0].value+timedelta(days=3)==cell.value:
						# print('dsakjddhjkahdsajh')
						if groupType=='blackout':
							cell.offset(column=4).value=',\n '.join(str(hotelName(elem)) for elem in hotelDict)
							print(cell.offset(column=4).value)
						if groupType=='one':
							cell.offset(column=5).value=',\n '.join(str(hotelName(elem)) for elem in hotelDict)
							print(cell.offset(column=5).value)							
						if groupType=='two':
							cell.offset(column=6).value=',\n '.join(str(hotelName(elem)) for elem in hotelDict)
							print(cell.offset(column=6).value)							
						projectCounter=projectCounter+1
					# elif(pd.notnull(cell.value)):
					# 	if groupType=='blackout':
					# 		cell.offset(column=4).value="N/A"
					# 		print('nab')
					# 	if groupType=='one':
					# 		cell.offset(column=5).value="N/A"
					# 		print('nao')							
					# 	if groupType=='two':
					# 		cell.offset(column=6).value="N/A"														
					# 		print('nat')
	wbProjects.save('F:/Steve/Testing Projects.xlsx')

# for dateName,dateRangeRow in dateRangesDF.iteritems():
for name, values in hotelBlackoutDF.iteritems():
	blackout= {}
	for x in range(0, 16):
		if(values[x]=='Blackout dates'):
			blackout[x]=hotelName(x)
			ws[counter][2].value=',\n '.join(str(hotelName(elem)) for elem in blackout)
			wb.save('F:/Steve/Testing Hotels.xlsx')
			if counter>1:
				projectAndHotel(blackout, counter, projectCounter,'blackout')
			wbProjects.save('F:/Steve/Testing Projects.xlsx')
			# print("Counted")
			# print(val)
	if counter>40:
		counter=1
	else:
		counter=counter+1
	# print(counter)
counter=1

for name, values in hotelOneDF.iteritems():
	one= {}
	for x in range(0, 16):
		if(values[x]=='Hotel can accommodate ONE (1) '):
			one[x]=hotelName(x)
			ws[counter][3].value=',\n '.join(str(hotelName(elem)) for elem in one)
			if counter>1:
				projectAndHotel(one, counter, projectCounter,'one')			
			wb.save('F:/Steve/Testing Hotels.xlsx')
			# print("Counted")
			# print(val)
	if counter>40:
		counter=1
	else:
		counter=counter+1
	# print(counter)

counter=1
for name, values in hotelTwoDF.iteritems():
	two= {}
	for x in range(0, 16):
		if(values[x]=='Hotel CAN accommodate two (2) separate IVLP projects during this week'):
			two[x]=hotelName(x)
			ws[counter][4].value=',\n '.join(str(hotelName(elem)) for elem in two)
			if counter>1:
				projectAndHotel(two, counter, projectCounter,'two')			
			wb.save('F:/Steve/Testing Hotels.xlsx')
			# print("Counted")
			# print(val)
	if counter>40:
		counter=1
	else:
		counter=counter+1
	# print(counter)


wb.save('F:/Steve/Testing Hotels.xlsx')


